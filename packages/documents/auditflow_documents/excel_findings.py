from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_ALIASES = {
    "title": {"title", "finding", "finding title", "name"},
    "risk_rating": {"risk", "risk rating", "severity", "priority"},
    "condition": {"condition", "observation", "issue"},
    "recommendation": {"recommendation", "recommended action", "action"},
    "owner": {"owner", "management owner", "responsible party"},
    "due_date": {"due date", "target date", "remediation date"},
    "status": {"status", "state"},
}


@dataclass(frozen=True)
class FindingRecord:
    row_number: int
    title: str
    risk_rating: str | None
    condition: str | None
    recommendation: str | None
    owner: str | None
    due_date: str | None
    status: str | None
    source_sheet: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def ingest_findings_workbook(path: str | Path, sheet_name: str | None = None) -> list[FindingRecord]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_map = _build_header_map(rows[0])
    findings: list[FindingRecord] = []
    for offset, row in enumerate(rows[1:], start=2):
        title = _cell(row, header_map.get("title"))
        if not title:
            continue
        findings.append(
            FindingRecord(
                row_number=offset,
                title=title,
                risk_rating=_cell(row, header_map.get("risk_rating")),
                condition=_cell(row, header_map.get("condition")),
                recommendation=_cell(row, header_map.get("recommendation")),
                owner=_cell(row, header_map.get("owner")),
                due_date=_cell(row, header_map.get("due_date")),
                status=_cell(row, header_map.get("status")),
                source_sheet=sheet.title,
            )
        )
    return findings


def _build_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    normalized = {
        _normalize(value): index
        for index, value in enumerate(header_row)
        if value is not None and _normalize(value)
    }
    result: dict[str, int] = {}
    for field, aliases in REQUIRED_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                result[field] = normalized[alias]
                break
    if "title" not in result:
        raise ValueError("Findings workbook must include a finding title column.")
    return result


def _cell(row: tuple[Any, ...], index: int | None) -> str | None:
    if index is None or index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    return str(value).strip()


def _normalize(value: Any) -> str:
    return str(value).strip().lower().replace("_", " ")

