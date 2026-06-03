from __future__ import annotations

import hashlib
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from auditflow_pptx import collect_metrics


MANIFEST_VERSION = "2026-06-03.1"


@dataclass(frozen=True)
class FileMetadata:
    path: str
    filename: str
    sha256: str
    size_bytes: int
    kind: str
    details: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class OperationRecord:
    operation_type: str
    target: str
    details: dict[str, Any]
    confidence: float | None = None
    fallback_used: bool = False
    warnings: list[str] | None = None

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["warnings"] = self.warnings or []
        return data


@dataclass(frozen=True)
class OperationManifest:
    manifest_version: str
    generated_at: str
    source_deck: FileMetadata
    findings_workbook: FileMetadata
    output_deck: FileMetadata
    slides_modified: list[int]
    slides_duplicated: list[dict[str, Any]]
    tables_updated: list[dict[str, Any]]
    charts_updated: list[dict[str, Any]]
    placeholders_mapped: list[dict[str, Any]]
    operations: list[OperationRecord]
    validation_results: dict[str, Any]
    preservation_benchmark_results: dict[str, Any]
    warnings: list[str]
    fallback_behavior_used: list[str]

    def to_dict(self) -> dict[str, Any]:
        return {
            "manifest_version": self.manifest_version,
            "generated_at": self.generated_at,
            "source_deck": self.source_deck.to_dict(),
            "findings_workbook": self.findings_workbook.to_dict(),
            "output_deck": self.output_deck.to_dict(),
            "slides_modified": self.slides_modified,
            "slides_duplicated": self.slides_duplicated,
            "tables_updated": self.tables_updated,
            "charts_updated": self.charts_updated,
            "placeholders_mapped": self.placeholders_mapped,
            "operations": [operation.to_dict() for operation in self.operations],
            "validation_results": self.validation_results,
            "preservation_benchmark_results": self.preservation_benchmark_results,
            "warnings": self.warnings,
            "fallback_behavior_used": self.fallback_behavior_used,
        }


def file_metadata(path: str | Path, kind: str) -> FileMetadata:
    file_path = Path(path)
    details: dict[str, Any] = {}
    if kind == "pptx":
        metrics = collect_metrics(file_path)
        details["pptx_metrics"] = metrics.to_dict()
    elif kind == "xlsx":
        details["workbook"] = workbook_metadata(file_path)
    return FileMetadata(
        path=str(file_path),
        filename=file_path.name,
        sha256=_sha256(file_path),
        size_bytes=file_path.stat().st_size,
        kind=kind,
        details=details,
    )


def workbook_metadata(path: str | Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        sheets.append({
            "name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
        })
    return {"sheet_count": len(sheets), "sheets": sheets}


def build_preservation_benchmark_snapshot(preservation: dict[str, Any], package_validation: dict[str, Any]) -> dict[str, Any]:
    checks = preservation.get("checks", {})
    package_valid = bool(package_validation.get("valid"))
    dimensions = {
        "layout_preservation": _score(checks.get("layouts_preserved")),
        "notes_preservation": _score(checks.get("notes_preserved")),
        "chart_preservation": _score(checks.get("charts_preserved")),
        "table_preservation": _score(checks.get("tables_preserved")),
        "theme_preservation": _score(checks.get("themes_preserved")),
        "editability": _score(checks.get("editable_text_present")),
        "corruption_rate": _score(package_valid),
        "export_integrity": _score(package_valid and not package_validation.get("issues")),
    }
    return {
        "score": round(sum(dimensions.values()) / len(dimensions), 3),
        "dimensions": dimensions,
    }


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _score(value: Any) -> float:
    return 1.0 if bool(value) else 0.0

